import streamlit as st
import pandas as pd
import numpy as np
import openpyxl

st.title("🔎 Detector de Erros em Planilhas")

uploaded_file = st.file_uploader(
    "Faça upload da planilha", type=["xlsx"]
)

if uploaded_file:

    # ── Leitura do arquivo ──────────────────────────────────────
    df = pd.read_excel(uploaded_file)

    st.subheader("Visualização dos dados")
    st.dataframe(df)

    # ── Duplicados ──────────────────────────────────────────────
    duplicados = df[df.duplicated()]

    st.subheader("Valores duplicados")
    st.write("Linhas com valores duplicados:", duplicados.shape[0])
    st.dataframe(duplicados)

    st.download_button(
        "Baixar duplicados",
        duplicados.to_csv(index=False),
        "duplicados.csv",
        "text/csv"
    )

    # ── Valores nulos ───────────────────────────────────────────
    st.subheader("Valores nulos")

    nulls = df[df.isnull().any(axis=1)]

    st.write("Linhas com valores nulos:", nulls.shape[0])
    st.dataframe(nulls)

    st.download_button(
        "Baixar valores nulos",
        nulls.to_csv(index=False),
        "valores_nulos.csv",
        "text/csv"
    )

    # ── Outliers ────────────────────────────────────────────────
    st.subheader("Outliers")

    colunas_numericas = df.select_dtypes(include=np.number).columns

    coluna = st.selectbox(
        "Escolha a coluna para detectar outliers",
        colunas_numericas
    )

    Q1 = df[coluna].quantile(0.25)
    Q3 = df[coluna].quantile(0.75)
    IQR = Q3 - Q1
    limite_inferior = Q1 - 1.5 * IQR
    limite_superior = Q3 + 1.5 * IQR

    outliers = df[
        (df[coluna] < limite_inferior) |
        (df[coluna] > limite_superior)
    ]

    st.write("Outliers encontrados:", outliers.shape[0])
    st.dataframe(outliers)

    st.download_button(
        "Baixar Outliers",
        outliers.to_csv(index=False),
        "outliers.csv",
        "text/csv"
    )

    # ── Detector de fórmulas ausentes ───────────────────────────
    st.subheader("🔍 Células sem fórmula")

    colunas_disponiveis = ["Selecione uma coluna..."] + df.columns.tolist()

    coluna_verificar = st.selectbox(
        "Escolha a coluna que deveria ter fórmula em todas as linhas",
        colunas_disponiveis,
        key="coluna_formula"
    )

    if coluna_verificar != "Selecione uma coluna...":
        try:
            uploaded_file.seek(0)
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active

            cabecalhos = [
                ws.cell(row=1, column=c).value
                for c in range(1, ws.max_column + 1)
            ]

            if coluna_verificar in cabecalhos:
                col_idx = cabecalhos.index(coluna_verificar) + 1

                resultados = []
                for linha in range(2, ws.max_row + 1):
                    celula = ws.cell(row=linha, column=col_idx)
                    valor = celula.value

                    if valor is None:
                        status = "⚠️ VAZIA"
                    elif isinstance(valor, str) and valor.startswith("="):
                        status = "✅ COM FÓRMULA"
                    else:
                        status = "❌ SEM FÓRMULA"

                    resultados.append({
                        "Linha (Excel)": linha,
                        "Valor encontrado": valor,
                        "Status": status
                    })

                df_resultado = pd.DataFrame(resultados)

                com_formula = (df_resultado["Status"] == "✅ COM FÓRMULA").sum()
                sem_formula = (df_resultado["Status"] == "❌ SEM FÓRMULA").sum()
                vazias      = (df_resultado["Status"] == "⚠️ VAZIA").sum()

                col1, col2, col3 = st.columns(3)
                col1.metric("✅ Com fórmula", com_formula)
                col2.metric("❌ Sem fórmula", sem_formula)
                col3.metric("⚠️ Vazias", vazias)

                df_problemas = df_resultado[
                    df_resultado["Status"] != "✅ COM FÓRMULA"
                ]

                st.write("Linhas com problemas:", df_problemas.shape[0])
                st.dataframe(df_problemas, use_container_width=True)

                if not df_problemas.empty:
                    st.download_button(
                        "Baixar células problemáticas",
                        df_problemas.to_csv(index=False),
                        "celulas_sem_formula.csv",
                        "text/csv"
                    )
                else:
                    st.success(
                        "Todas as células da coluna possuem fórmula! 🎉"
                    )

        except Exception as e:
            st.error(f"Erro ao processar o arquivo: {e}")

            #py -m streamlit run TccConf4.py